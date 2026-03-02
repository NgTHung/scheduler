"""
Orientation Event Scheduler — Web UI (Streamlit)
=================================================
Launch with:
    streamlit run app.py

Features:
  • Upload Excel files (combined or per-role, checkbox or text format)
  • Live-editable data tables for hosts, mentors, students — grouped by day
  • Shift label mapping (Ca 1 → 8h00 - 8h50) — editable on the fly
  • One-click ILP solve with schedule display
  • Timetable split by role & day
  • Download results as JSON or Excel
"""

from __future__ import annotations

import io
import json
import re
import sys
from collections import Counter, OrderedDict, defaultdict
from pathlib import Path
from typing import Any

import pandas as pd
import streamlit as st

sys.path.insert(0, str(Path(__file__).resolve().parent))

from models import Host, Mentor, Student, ScheduledSession
from solver import solve
from convert_excel import (
    parse_workbook,
    parse_combined_workbook,
    _normalize_day_label,
)

st.set_page_config(
    page_title="Orientation Scheduler",
    page_icon="📅",
    layout="wide",
    initial_sidebar_state="expanded",
)

DEFAULT_SHIFT_LABELS: dict[int, str] = {
    1:  "Ca 1  (8h00 - 8h50)",
    2:  "Ca 2  (9h00 - 9h50)",
    3:  "Ca 3  (10h00 - 10h50)",
    4:  "Ca 4  (11h00 - 11h50)",
    5:  "Ca 5  (13h00 - 13h50)",
    6:  "Ca 6  (14h00 - 14h50)",
    7:  "Ca 7  (15h00 - 15h50)",
    8:  "Ca 8  (16h00 - 16h50)",
    9:  "Ca 9  (17h00 - 17h50)",
    10: "Ca 10 (18h00 - 18h50)",
    11: "Ca 11 (19h00 - 19h50)",
    12: "Ca 12 (20h00 - 20h50)",
}

def _default(key: str, value: Any):
    if key not in st.session_state:
        st.session_state[key] = value


_default("hosts_data", [])
_default("mentors_data", [])
_default("students_data", [])
_default("time_slots", [])
_default("schedule_result", None)
_default("data_loaded", False)
_default("shift_labels", dict(DEFAULT_SHIFT_LABELS))
_default("widget_rev", 0)  # increment to force fresh widget keys after clear/reload
_default("pending_shift_sync", None)  # dict describing pending add/remove shifts


def _clear_all_data():
    """Wipe all loaded data and solver results, reset widget keys."""
    st.session_state.hosts_data = []
    st.session_state.mentors_data = []
    st.session_state.students_data = []
    st.session_state.time_slots = []
    st.session_state.schedule_result = None
    st.session_state.data_loaded = False
    st.session_state.solver_error = None
    st.session_state.widget_rev += 1


def _input_to_json_bytes() -> bytes:
    """Serialize current input data (no schedule) to JSON."""
    labels = st.session_state.get("shift_labels", {})
    output = {
        "time_slots": st.session_state.time_slots,
        "shift_labels": {str(k): v for k, v in labels.items()},
        "hosts": st.session_state.hosts_data,
        "mentors": st.session_state.mentors_data,
        "students": st.session_state.students_data,
    }
    return json.dumps(output, indent=2, ensure_ascii=False).encode("utf-8")


def _input_to_excel_bytes() -> bytes:
    """Serialize current input data to Excel (one sheet per role)."""
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        time_slots = st.session_state.time_slots
        for role, data_key, major_key in [
            ("Hosts", "hosts_data", None),
            ("Mentors", "mentors_data", "major"),
            ("Students", "students_data", "desired_major"),
        ]:
            people = st.session_state[data_key]
            rows = []
            for p in people:
                avail_set = set(p.get("available_slots", []))
                row: dict[str, Any] = {"Name": p["name"]}
                if major_key:
                    row["Major"] = p.get(major_key, "")
                for slot in time_slots:
                    row[_full_slot_display(slot)] = "✓" if slot in avail_set else ""
                rows.append(row)
            pd.DataFrame(rows).to_excel(writer, sheet_name=role, index=False)
    return buf.getvalue()

def _parse_slot(slot_id: str) -> tuple[str, int]:
    """Split '13/6_3' → ('13/6', 3)."""
    parts = slot_id.rsplit("_", 1)
    if len(parts) == 2 and parts[1].isdigit():
        return parts[0], int(parts[1])
    return slot_id, 0


def _group_slots_by_day(time_slots: list[str]) -> OrderedDict[str, list[str]]:
    """Group time slots by their day portion, preserving order."""
    groups: OrderedDict[str, list[str]] = OrderedDict()
    for slot in time_slots:
        day, _ = _parse_slot(slot)
        groups.setdefault(day, []).append(slot)
    return groups


def _slot_col_header(slot_id: str) -> str:
    """Short column header for a slot within a day table."""
    _, shift_num = _parse_slot(slot_id)
    labels: dict[int, str] = st.session_state.get("shift_labels", {})
    label = labels.get(shift_num)
    if label:
        return label
    return f"Shift {shift_num}"


def _full_slot_display(slot_id: str) -> str:
    """Full display with day + shift label."""
    day, shift_num = _parse_slot(slot_id)
    labels: dict[int, str] = st.session_state.get("shift_labels", {})
    label = labels.get(shift_num)
    if label:
        return f"{day} — {label}"
    return f"{day} — Shift {shift_num}"

def _save_uploaded_file(uploaded) -> Path:
    tmp = Path("_tmp_upload") / uploaded.name
    tmp.parent.mkdir(exist_ok=True)
    tmp.write_bytes(uploaded.getvalue())
    return tmp

def _people_to_day_df(
    people: list[dict], role: str, day_slots: list[str],
) -> pd.DataFrame:
    """
    Build a DataFrame for ONE day's slots only.
    Columns: Name, [Major], shift1_col, shift2_col, ...
    """
    rows = []
    for p in people:
        avail_set = set(p.get("available_slots", []))
        row: dict[str, Any] = {"Name": p["name"]}
        if role in ("mentor", "student"):
            row["Major"] = p.get("major", "") if role == "mentor" else p.get("desired_major", "")
        for slot in day_slots:
            row[slot] = slot in avail_set
        rows.append(row)

    if rows:
        return pd.DataFrame(rows)

    # Empty frame
    cols: dict[str, Any] = {"Name": pd.Series(dtype="str")}
    if role in ("mentor", "student"):
        cols["Major"] = pd.Series(dtype="str")
    for slot in day_slots:
        cols[slot] = pd.Series(dtype="bool")
    return pd.DataFrame(cols)


def _person_key(p: dict, role: str) -> tuple[str, str]:
    """Unique key for a person: (name, major). Preserves same-name people with different majors."""
    name = p.get("name", "")
    if role == "mentor":
        return (name, p.get("major", ""))
    elif role == "student":
        return (name, p.get("desired_major", ""))
    return (name, "")


def _sync_people_from_day_dfs(
    role: str,
    days: OrderedDict[str, list[str]],
    edited_dfs: dict[str, pd.DataFrame],
    current_people: list[dict],
) -> list[dict]:
    """
    Merge edited per-day DataFrames back into the people list.
    Handles new rows added in any day tab.
    Uses (name, major) as key to preserve duplicate names with different majors.
    """
    # Collect all known people (preserving original data as base)
    people_map: dict[tuple[str, str], dict] = {}
    for p in current_people:
        pk = _person_key(p, role)
        people_map[pk] = {
            "name": p["name"],
            "available_slots": set(p.get("available_slots", [])),
        }
        if role == "mentor":
            people_map[pk]["major"] = p.get("major", "")
        elif role == "student":
            people_map[pk]["desired_major"] = p.get("desired_major", "")

    # Merge in edits from each day tab
    for day_label, day_slots in days.items():
        df = edited_dfs.get(day_label)
        if df is None:
            continue

        for _, row in df.iterrows():
            name = str(row.get("Name", "")).strip()
            if not name:
                continue

            major_str = str(row.get("Major", "")).strip() if "Major" in row.index else ""
            pk = (name, major_str) if role in ("mentor", "student") else (name, "")

            if pk not in people_map:
                people_map[pk] = {"name": name, "available_slots": set()}
                if role == "mentor":
                    people_map[pk]["major"] = major_str
                elif role == "student":
                    people_map[pk]["desired_major"] = major_str

            # Update major if present
            if role == "mentor" and "Major" in row.index:
                people_map[pk]["major"] = major_str
            elif role == "student" and "Major" in row.index:
                people_map[pk]["desired_major"] = major_str

            # Remove old slots for this day, then add checked ones
            avail: set = people_map[pk]["available_slots"]
            for s in day_slots:
                avail.discard(s)
            for s in day_slots:
                if row.get(s, False):
                    avail.add(s)

    # Convert back to list[dict] — sort slots to canonical order
    all_slots = []
    for sl in days.values():
        all_slots.extend(sl)

    result = []
    for p in people_map.values():
        entry = dict(p)
        entry["available_slots"] = [s for s in all_slots if s in p["available_slots"]]
        result.append(entry)
    return result


def _build_model_objects(
    hosts_data: list[dict],
    mentors_data: list[dict],
    students_data: list[dict],
):
    hosts = [Host(name=h["name"], available_slots=h["available_slots"]) for h in hosts_data]
    mentors = [Mentor(name=m["name"], major=m["major"], available_slots=m["available_slots"]) for m in mentors_data]
    students = [Student(name=s["name"], desired_major=s["desired_major"], available_slots=s["available_slots"]) for s in students_data]
    return hosts, mentors, students

def _sessions_to_df(sessions: list[ScheduledSession], use_labels: bool = True) -> pd.DataFrame:
    rows = []
    for s in sessions:
        rows.append({
            "Time Slot": _full_slot_display(s.time_slot) if use_labels else s.time_slot,
            "Host": s.host,
            "Mentor": s.mentor,
            "Student": s.student,
            "Major": s.major,
        })
    return pd.DataFrame(rows)


def _build_role_day_timetable(
    sessions: list[ScheduledSession],
    role: str,
    day_label: str,
    day_slots: list[str],
) -> pd.DataFrame:
    """
    Build timetable for ONE role on ONE day.
    Rows = people of that role. Columns = shifts.
    Cells = paired info (e.g. mentor sees student name, host sees mentor+student).
    """
    day_sessions = [s for s in sessions if _parse_slot(s.time_slot)[0] == day_label]

    people_data: dict[str, dict[str, str]] = {}

    for s in day_sessions:
        if role == "host":
            person = s.host
            cell = f"{s.mentor} + {s.student} ({s.major})"
        elif role == "mentor":
            person = s.mentor
            cell = f"{s.student} | Host: {s.host}"
        else:
            person = s.student
            cell = f"{s.mentor} | Host: {s.host}"

        if person not in people_data:
            people_data[person] = {}
        people_data[person][s.time_slot] = cell

    if not people_data:
        return pd.DataFrame()

    rows = []
    for name in sorted(people_data):
        row: dict[str, str] = {"Name": name}
        for slot in day_slots:
            header = _slot_col_header(slot)
            row[header] = people_data[name].get(slot, "")
        rows.append(row)

    return pd.DataFrame(rows)


def _constraint_check(
    sessions: list[ScheduledSession],
    mentors: list[Mentor],
) -> tuple[bool, list[str]]:
    msgs: list[str] = []
    ok = True

    host_slots: dict[tuple, list] = defaultdict(list)
    mentor_slots: dict[tuple, list] = defaultdict(list)
    student_slots: dict[tuple, list] = defaultdict(list)

    for s in sessions:
        host_slots[(s.time_slot, s.host)].append(s)
        mentor_slots[(s.time_slot, s.mentor)].append(s)
        student_slots[(s.time_slot, s.student)].append(s)

    for key, lst in host_slots.items():
        if len(lst) > 1:
            msgs.append(f"FAIL: Host **{key[1]}** double-booked at {_full_slot_display(key[0])}")
            ok = False
    for key, lst in mentor_slots.items():
        if len(lst) > 1:
            msgs.append(f"FAIL: Mentor **{key[1]}** double-booked at {_full_slot_display(key[0])}")
            ok = False
    for key, lst in student_slots.items():
        if len(lst) > 1:
            msgs.append(f"FAIL: Student **{key[1]}** double-booked at {_full_slot_display(key[0])}")
            ok = False

    mentor_names = {m.name for m in mentors}
    scheduled_mentors = {s.mentor for s in sessions}
    missing = mentor_names - scheduled_mentors
    if missing:
        for mn in sorted(missing):
            msgs.append(f"FAIL: Mentor **{mn}** has 0 sessions")
        ok = False

    if ok:
        msgs.append("ALL CONSTRAINTS SATISFIED")
    return ok, msgs


def _result_to_json_bytes(
    sessions: list[ScheduledSession],
    time_slots: list[str],
    hosts_data: list[dict],
    mentors_data: list[dict],
    students_data: list[dict],
) -> bytes:
    labels = st.session_state.get("shift_labels", {})
    output = {
        "time_slots": time_slots,
        "shift_labels": {str(k): v for k, v in labels.items()},
        "hosts": hosts_data,
        "mentors": mentors_data,
        "students": students_data,
        "schedule": [
            {
                "time_slot": s.time_slot,
                "time_label": _full_slot_display(s.time_slot),
                "host": s.host,
                "mentor": s.mentor,
                "student": s.student,
                "major": s.major,
            }
            for s in sessions
        ],
    }
    return json.dumps(output, indent=2, ensure_ascii=False).encode("utf-8")


def _build_mentor_day_timetable(
    sessions: list[ScheduledSession],
    day_label: str,
    day_slots: list[str],
) -> pd.DataFrame:
    """
    Build mentor timetable for ONE day.
    Rows = mentors, Columns = shifts.
    Cells = "Student | Host: host (Major)".
    """
    day_sessions = [s for s in sessions if _parse_slot(s.time_slot)[0] == day_label]
    people_data: dict[str, dict[str, str]] = {}

    for s in day_sessions:
        person = s.mentor
        cell = f"{s.student} | Host: {s.host} ({s.major})"
        if person not in people_data:
            people_data[person] = {}
        people_data[person][s.time_slot] = cell

    if not people_data:
        return pd.DataFrame()

    rows = []
    for name in sorted(people_data):
        row: dict[str, str] = {"Tên CVHN": name}
        for slot in day_slots:
            header = _slot_col_header(slot)
            row[header] = people_data[name].get(slot, "")
        rows.append(row)

    return pd.DataFrame(rows)


# ---- Major color palette for summary sheet -------------------------------- #
_MAJOR_COLORS: list[str] = [
    "DAEEF3",  # light blue
    "D5E8D4",  # light green
    "FFF2CC",  # light yellow
    "F8CECC",  # light red/pink
    "E1D5E7",  # light purple
    "FFE6CC",  # light orange
]


def _result_to_excel_bytes(sessions: list[ScheduledSession]) -> bytes:
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    all_slots_in_result = sorted(
        set(s.time_slot for s in sessions),
        key=lambda x: st.session_state.time_slots.index(x)
        if x in st.session_state.time_slots else 0,
    )
    days = _group_slots_by_day(all_slots_in_result)

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        # ---- Per-day mentor timetables ------------------------------------ #
        for day_label, day_slots in days.items():
            tt = _build_mentor_day_timetable(sessions, day_label, day_slots)
            if tt.empty:
                continue
            safe_day = day_label.replace("/", "-")
            sheet_name = f"Ngày {safe_day}"[:31]
            tt.to_excel(writer, sheet_name=sheet_name, index=False)

        # ---- Summary tab (grouped by major) ------------------------------- #
        summary_rows = []
        for s in sessions:
            day, shift = _parse_slot(s.time_slot)
            summary_rows.append({
                "Ngành": s.major,
                "Tên CVHN": s.mentor,
                "Tên Host": s.host,
                "Ngày": day,
                "Ca": shift,
                "NTG": s.student,
            })
        if not summary_rows:
            return buf.getvalue()

        df_summary = pd.DataFrame(summary_rows)
        # Sort by major then day/shift for clean grouping
        df_summary = df_summary.sort_values(
            by=["Ngành", "Ngày", "Ca"],
            key=lambda col: col if col.name != "Ngày" else col.map(
                lambda d: st.session_state.time_slots.index(
                    next((s for s in st.session_state.time_slots if s.startswith(d + "_")), d)
                ) if any(s.startswith(d + "_") for s in st.session_state.time_slots) else 0
            ),
        ).reset_index(drop=True)

        sheet_name = "Tổng hợp"
        df_summary.to_excel(writer, sheet_name=sheet_name, index=False)

        # ---- Style the summary sheet -------------------------------------- #
        ws = writer.sheets[sheet_name]

        # Header style
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        # Assign colours to majors in order of appearance
        major_col_idx = 1  # Column A = Ngành
        seen_majors: list[str] = []
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=major_col_idx, max_col=major_col_idx):
            val = row[0].value
            if val and val not in seen_majors:
                seen_majors.append(val)
        major_color_map: dict[str, str] = {}
        for i, mj in enumerate(seen_majors):
            major_color_map[mj] = _MAJOR_COLORS[i % len(_MAJOR_COLORS)]

        # Apply row colours and borders
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            major_val = row[0].value
            fill_color = major_color_map.get(major_val, "FFFFFF")
            fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            for cell in row:
                cell.fill = fill
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")

        # Merge cells in the Ngành column for same major
        start_row = 2
        prev_major = ws.cell(row=2, column=1).value
        for r in range(3, ws.max_row + 2):  # +2 to process last group
            current = ws.cell(row=r, column=1).value if r <= ws.max_row else None
            if current != prev_major:
                if r - 1 > start_row:
                    ws.merge_cells(
                        start_row=start_row, start_column=1,
                        end_row=r - 1, end_column=1,
                    )
                    ws.cell(row=start_row, column=1).alignment = Alignment(
                        horizontal="center", vertical="center", wrap_text=True,
                    )
                start_row = r
                prev_major = current

        # Auto-fit column widths
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 4, 35)

    return buf.getvalue()

def _render_sidebar():
    with st.sidebar:
        st.header("Data Import")

        mode = st.radio(
            "Input mode",
            ["Combined workbook", "Separate files (per role)", "Hybrid (combined + overrides)", "Manual entry"],
            help="Choose how to load your data",
        )

        if mode == "Combined workbook":
            uploaded = st.file_uploader(
                "Upload combined Excel",
                type=["xlsx", "xls"],
                key="combined_upload",
            )
            if uploaded and st.button("Load", key="load_combined", type="primary"):
                _clear_all_data()
                _load_combined(uploaded)

        elif mode == "Separate files (per role)":
            hosts_file = st.file_uploader("Hosts Excel", type=["xlsx", "xls"], key="hosts_upload")
            mentors_file = st.file_uploader("Mentors Excel", type=["xlsx", "xls"], key="mentors_upload")
            students_file = st.file_uploader("Students Excel", type=["xlsx", "xls"], key="students_upload")
            if st.button("Load", key="load_separate", type="primary"):
                _clear_all_data()
                _load_separate(hosts_file, mentors_file, students_file)

        elif mode == "Hybrid (combined + overrides)":
            st.caption(
                "Upload a combined workbook as base, then optionally override "
                "individual roles with separate files."
            )
            combined_file = st.file_uploader(
                "Combined base workbook",
                type=["xlsx", "xls"],
                key="hybrid_combined_upload",
            )
            st.caption("Override (leave empty to use combined):")
            hosts_override = st.file_uploader("Hosts override", type=["xlsx", "xls"], key="hybrid_hosts")
            mentors_override = st.file_uploader("Mentors override", type=["xlsx", "xls"], key="hybrid_mentors")
            students_override = st.file_uploader("Students override", type=["xlsx", "xls"], key="hybrid_students")
            if combined_file and st.button("Load", key="load_hybrid", type="primary"):
                _clear_all_data()
                _load_hybrid(combined_file, hosts_override, mentors_override, students_override)

        elif mode == "Manual entry":
            st.caption("Configure time slots, then add people in the main area.")
            slots_text = st.text_area(
                "Time slots (one per line)",
                value="\n".join(st.session_state.time_slots),
                height=150,
                help="e.g. 13/6_1, 13/6_2, 14/6_1, ...",
            )
            if st.button("Apply slots", key="apply_slots"):
                slots = [s.strip() for s in slots_text.strip().split("\n") if s.strip()]
                st.session_state.time_slots = slots
                st.session_state.data_loaded = True
                st.rerun()

        st.divider()
        if st.session_state.data_loaded:
            if st.button("Clear All Data", key="clear_all", type="secondary"):
                _clear_all_data()
                st.rerun()
        if st.session_state.data_loaded:
            st.subheader("Data Summary")
            st.metric("Time Slots", len(st.session_state.time_slots))
            col1, col2, col3 = st.columns(3)
            col1.metric("Hosts", len(st.session_state.hosts_data))
            col2.metric("Mentors", len(st.session_state.mentors_data))
            col3.metric("Students", len(st.session_state.students_data))

            majors = sorted({m.get("major", "?") for m in st.session_state.mentors_data})
            if majors:
                st.caption(f"Majors: {', '.join(majors)}")

            for label, people in [
                ("Hosts", st.session_state.hosts_data),
                ("Mentors", st.session_state.mentors_data),
                ("Students", st.session_state.students_data),
            ]:
                no_avail = [p["name"] for p in people if not p.get("available_slots")]
                if no_avail:
                    st.warning(f"{label} with 0 availability: {', '.join(no_avail)}")


def _load_combined(uploaded):
    with st.spinner("Parsing combined workbook..."):
        tmp = _save_uploaded_file(uploaded)
        try:
            hosts, mentors, students, slots = parse_combined_workbook(str(tmp))
            st.session_state.hosts_data = hosts
            st.session_state.mentors_data = mentors
            st.session_state.students_data = students
            st.session_state.time_slots = slots
            st.session_state.data_loaded = True
            st.session_state.schedule_result = None
            st.success(f"Loaded {len(hosts)} hosts, {len(mentors)} mentors, {len(students)} students")
        except Exception as e:
            st.error(f"Error parsing file: {e}")
        finally:
            tmp.unlink(missing_ok=True)


def _load_separate(hosts_file, mentors_file, students_file):
    if not all([hosts_file, mentors_file, students_file]):
        st.warning("Please upload all three files.")
        return
    with st.spinner("Parsing workbooks..."):
        paths = {}
        try:
            for label, f, role in [
                ("hosts", hosts_file, "host"),
                ("mentors", mentors_file, "mentor"),
                ("students", students_file, "student"),
            ]:
                paths[label] = _save_uploaded_file(f)

            h_entries, h_slots = parse_workbook(str(paths["hosts"]), "host")
            m_entries, m_slots = parse_workbook(str(paths["mentors"]), "mentor")
            s_entries, s_slots = parse_workbook(str(paths["students"]), "student")

            seen: set[str] = set()
            all_slots: list[str] = []
            for slot_list in [h_slots, m_slots, s_slots]:
                for s in slot_list:
                    if s not in seen:
                        seen.add(s)
                        all_slots.append(s)

            st.session_state.hosts_data = h_entries
            st.session_state.mentors_data = m_entries
            st.session_state.students_data = s_entries
            st.session_state.time_slots = all_slots
            st.session_state.data_loaded = True
            st.session_state.schedule_result = None
            st.success(f"Loaded {len(h_entries)} hosts, {len(m_entries)} mentors, {len(s_entries)} students")
        except Exception as e:
            st.error(f"Error parsing files: {e}")
        finally:
            for p in paths.values():
                p.unlink(missing_ok=True)


def _load_hybrid(combined_file, hosts_override, mentors_override, students_override):
    """Load from combined workbook, then override individual roles if provided."""
    with st.spinner("Parsing hybrid sources..."):
        paths: list[Path] = []
        try:
            # 1) Parse combined base
            combined_tmp = _save_uploaded_file(combined_file)
            paths.append(combined_tmp)
            hosts, mentors, students, base_slots = parse_combined_workbook(str(combined_tmp))
            slot_lists: list[list[str]] = [base_slots]

            # 2) Override per-role if files provided
            overrides = [
                (hosts_override, "host", "hosts"),
                (mentors_override, "mentor", "mentors"),
                (students_override, "student", "students"),
            ]
            for uploaded, role, label in overrides:
                if uploaded is None:
                    continue
                tmp = _save_uploaded_file(uploaded)
                paths.append(tmp)
                entries, slots = parse_workbook(str(tmp), role)
                slot_lists.append(slots)
                if role == "host":
                    hosts = entries
                elif role == "mentor":
                    mentors = entries
                else:
                    students = entries

            # 3) Merge time slots
            seen: set[str] = set()
            all_slots: list[str] = []
            for slot_list in slot_lists:
                for s in slot_list:
                    if s not in seen:
                        seen.add(s)
                        all_slots.append(s)

            st.session_state.hosts_data = hosts
            st.session_state.mentors_data = mentors
            st.session_state.students_data = students
            st.session_state.time_slots = all_slots
            st.session_state.data_loaded = True
            st.session_state.schedule_result = None

            sources = ["combined base"]
            for uploaded, _, label in overrides:
                if uploaded:
                    sources.append(f"{label} override")
            st.success(
                f"Loaded {len(hosts)} hosts, {len(mentors)} mentors, "
                f"{len(students)} students (from {', '.join(sources)})"
            )
        except Exception as e:
            st.error(f"Error parsing files: {e}")
        finally:
            for p in paths:
                p.unlink(missing_ok=True)

def _render_shift_label_editor():
    st.header("Shift Labels")
    st.caption(
        "Map shift numbers to readable names. These labels appear in column "
        "headers, timetables, and exports. Upload a JSON or edit the table below. "
        "Click **Apply Labels** to save your changes."
    )

    col_upload, col_reset = st.columns([3, 1])
    with col_upload:
        uploaded = st.file_uploader(
            "Upload label mapping (JSON)",
            type=["json"],
            key="label_json_upload",
            help='JSON object: {"1": "Ca 1 (8h00 - 8h50)", "2": "Ca 2 ...", ...}',
        )
        if uploaded:
            try:
                raw = json.loads(uploaded.getvalue().decode("utf-8"))
                st.session_state.shift_labels = {int(k): str(v) for k, v in raw.items()}
                st.session_state.pending_shift_sync = None
                st.session_state.widget_rev += 1
                st.success("Labels loaded from JSON!")
                st.rerun()
            except Exception as e:
                st.error(f"Invalid JSON: {e}")

    with col_reset:
        st.write("")
        st.write("")
        if st.button("Reset defaults", key="reset_labels"):
            st.session_state.shift_labels = dict(DEFAULT_SHIFT_LABELS)
            st.session_state.pending_shift_sync = None
            st.session_state.widget_rev += 1
            st.rerun()

    # Determine shifts present in data
    labels = st.session_state.shift_labels
    data_shifts: set[int] = set()
    if st.session_state.time_slots:
        for slot in st.session_state.time_slots:
            _, n = _parse_slot(slot)
            if n > 0:
                data_shifts.add(n)
    # Show at least shifts 1-12 or whatever is in labels/data
    all_shifts = sorted(set(range(1, 13)) | set(labels.keys()) | data_shifts)

    label_rows = []
    for i in all_shifts:
        label_rows.append({"Shift #": i, "Label": labels.get(i, f"Shift {i}")})

    label_df = pd.DataFrame(label_rows)

    rev = st.session_state.widget_rev
    edited_labels = st.data_editor(
        label_df,
        column_config={
            "Shift #": st.column_config.NumberColumn("Shift #", width="small"),
            "Label": st.column_config.TextColumn("Display Label", width="large"),
        },
        num_rows="dynamic",
        width='stretch',
        key=f"label_editor_v{rev}",
        hide_index=True,
    )
    editor_labels: dict[int, str] = {}
    for _, row in edited_labels.iterrows():
        shift_val = row.get("Shift #")
        if pd.isna(shift_val):
            continue
        shift = int(shift_val)
        label = str(row.get("Label", "")).strip()
        if shift > 0 and label:
            editor_labels[shift] = label
    has_label_changes = (editor_labels != labels)
    editor_nums = set(editor_labels.keys())
    removed = data_shifts - editor_nums
    added = editor_nums - data_shifts
    has_slot_changes = bool(removed or added) and bool(st.session_state.time_slots)
    if has_label_changes or has_slot_changes:
        btn_cols = st.columns([1, 4])
        with btn_cols[0]:
            apply_clicked = st.button("Apply Labels", key="apply_labels", type="primary")

        if has_slot_changes:
            parts = []
            if removed:
                parts.append(f"Remove shift(s): **{', '.join(str(s) for s in sorted(removed))}**")
            if added:
                parts.append(f"Add shift(s): **{', '.join(str(s) for s in sorted(added))}**")
            st.warning(
                "⚠️ Shift changes detected — " + "; ".join(parts)
                + ". Click **Apply Labels** to save & sync to data tables."
            )
        elif has_label_changes:
            st.info("Label changes detected. Click **Apply Labels** to save.")

        if apply_clicked:
            # Write labels to state
            st.session_state.shift_labels = dict(editor_labels)
            # Sync slot additions / removals
            if has_slot_changes:
                st.session_state.pending_shift_sync = {
                    "removed": sorted(removed),
                    "added": sorted(added),
                }
                _apply_shift_sync()
                st.session_state.pending_shift_sync = None
            # Refresh all widgets so editors pick up new columns / headers
            st.session_state.widget_rev += 1
            st.rerun()
    else:
        st.caption("Labels are up to date.")


def _apply_shift_sync():
    """Apply pending shift additions/removals to time_slots & availability."""
    pending = st.session_state.get("pending_shift_sync")
    if not pending:
        return

    existing_days = _group_slots_by_day(st.session_state.time_slots)
    removed_shifts = set(pending.get("removed", []))
    added_shifts = set(pending.get("added", []))

    # 1) Remove slots whose shift # was deleted
    if removed_shifts:
        removed_slots = {
            slot for slot in st.session_state.time_slots
            if _parse_slot(slot)[1] in removed_shifts
        }
        st.session_state.time_slots = [
            s for s in st.session_state.time_slots if s not in removed_slots
        ]
        for data_key in ("hosts_data", "mentors_data", "students_data"):
            for person in st.session_state[data_key]:
                person["available_slots"] = [
                    s for s in person.get("available_slots", [])
                    if s not in removed_slots
                ]

    # 2) Add slots for newly added shift #s (for every existing day)
    if added_shifts and existing_days:
        new_slots: list[str] = []
        for day_label in existing_days:
            for shift_num in sorted(added_shifts):
                slot_id = f"{day_label}_{shift_num}"
                if slot_id not in st.session_state.time_slots:
                    new_slots.append(slot_id)
        if new_slots:
            all_slots = st.session_state.time_slots + new_slots
            regrouped = _group_slots_by_day(all_slots)
            ordered: list[str] = []
            for day_label in regrouped:
                day_slots = regrouped[day_label]
                day_slots.sort(key=lambda s: _parse_slot(s)[1])
                ordered.extend(day_slots)
            st.session_state.time_slots = ordered

    # Force data editors to refresh with new columns
    st.session_state.widget_rev += 1

def _render_data_editor():
    st.header("Data Editor")
    st.caption("Edit availability per day. Add/remove people, toggle shifts, then run the solver.")

    time_slots = st.session_state.time_slots
    if not time_slots:
        st.info("No time slots loaded. Import data or enter time slots manually in the sidebar.")
        return
    save_col1, save_col2, save_spacer = st.columns([1, 1, 4])
    with save_col1:
        st.download_button(
            label="Save Input (JSON)",
            data=_input_to_json_bytes(),
            file_name="input_data.json",
            mime="application/json",
        )
    with save_col2:
        st.download_button(
            label="Save Input (Excel)",
            data=_input_to_excel_bytes(),
            file_name="input_data.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    days = _group_slots_by_day(time_slots)

    tab_hosts, tab_mentors, tab_students = st.tabs(["Hosts", "Mentors", "Students"])

    candidates: dict[str, tuple[str, list[dict], bool]] = {}  # role → (data_key, candidate, changed)

    with tab_hosts:
        candidates["host"] = _render_role_editor_by_day("host", days)
    with tab_mentors:
        candidates["mentor"] = _render_role_editor_by_day("mentor", days)
    with tab_students:
        candidates["student"] = _render_role_editor_by_day("student", days)
    any_changes = any(changed for _, _, changed in candidates.values())
    changed_roles = [role for role, (_, _, changed) in candidates.items() if changed]

    st.divider()
    btn_col, info_col = st.columns([1, 4])
    with btn_col:
        apply_clicked = st.button(
            "Apply Changes",
            key="apply_all_roles",
            type="primary" if any_changes else "secondary",
            disabled=not any_changes,
        )
    with info_col:
        if any_changes:
            st.info(
                f"Unsaved edits in **{', '.join(changed_roles)}** — "
                "click **Apply Changes** to commit all."
            )
        else:
            counts = []
            for role, (data_key, _, _) in candidates.items():
                people = st.session_state[data_key]
                n = len(people)
                n_avail = sum(1 for p in people if p.get("available_slots"))
                counts.append(f"{n} {role}s ({n_avail} with availability)")
            st.caption(" · ".join(counts))

    if apply_clicked:
        for role, (data_key, candidate, changed) in candidates.items():
            if changed:
                st.session_state[data_key] = candidate
        st.session_state.widget_rev += 1
        st.rerun()


def _render_role_editor_by_day(role: str, days: OrderedDict[str, list[str]]) -> tuple[str, list[dict], bool]:
    """Render day-tabbed editors for one role. Returns (data_key, candidate_people, has_changes)."""
    data_key = f"{role}s_data" if role != "host" else "hosts_data"
    people = st.session_state[data_key]

    day_labels = list(days.keys())
    if not day_labels:
        st.info("No days found in time slots.")
        return data_key, people, False

    # One tab per day
    rev = st.session_state.widget_rev
    day_tabs = st.tabs(list(day_labels))
    edited_dfs: dict[str, pd.DataFrame] = {}

    for day_tab, day_label in zip(day_tabs, day_labels):
        with day_tab:
            day_slots = days[day_label]
            df = _people_to_day_df(people, role, day_slots)

            col_config: dict[str, Any] = {
                "Name": st.column_config.TextColumn("Name", required=True, width="medium"),
            }
            if role in ("mentor", "student"):
                col_config["Major"] = st.column_config.TextColumn("Major", required=True, width="small")

            for slot in day_slots:
                header = _slot_col_header(slot)
                col_config[slot] = st.column_config.CheckboxColumn(
                    header, default=False, width="small",
                )

            edited_df = st.data_editor(
                df,
                column_config=col_config,
                num_rows="dynamic",
                width='stretch',
                key=f"editor_{role}_{day_label}_v{rev}",
                hide_index=True,
            )
            edited_dfs[day_label] = edited_df
    candidate = _sync_people_from_day_dfs(role, days, edited_dfs, people)
    has_changes = _people_differ(people, candidate)

    return data_key, candidate, has_changes


def _people_differ(old: list[dict], new: list[dict]) -> bool:
    """Quick check whether two people lists differ materially."""
    if len(old) != len(new):
        return True
    for a, b in zip(old, new):
        if a.get("name") != b.get("name"):
            return True
        if a.get("major") != b.get("major"):
            return True
        if a.get("desired_major") != b.get("desired_major"):
            return True
        if sorted(a.get("available_slots", [])) != sorted(b.get("available_slots", [])):
            return True
    return False

def _render_solver():
    st.header("Solver")

    col1, col2 = st.columns([1, 3])
    with col1:
        if st.button("Run Solver", type="primary", width='stretch'):
            _run_solver()
    with col2:
        if st.session_state.schedule_result is not None:
            sessions = st.session_state.schedule_result
            st.success(f"Solution found — **{len(sessions)} sessions** scheduled")
        elif st.session_state.get("solver_error"):
            st.error(st.session_state.solver_error)

    if st.session_state.schedule_result is not None:
        sessions = st.session_state.schedule_result
        hosts, mentors, students = _build_model_objects(
            st.session_state.hosts_data,
            st.session_state.mentors_data,
            st.session_state.students_data,
        )

        tab_schedule, tab_timetable, tab_summary, tab_export = st.tabs(
            ["Schedule", "Timetable", "Summary", "Export"]
        )

        with tab_schedule:
            _render_schedule_tab(sessions)
        with tab_timetable:
            _render_timetable_tab(sessions)
        with tab_summary:
            _render_summary_tab(sessions, mentors, students)
        with tab_export:
            _render_export_tab(sessions)


def _run_solver():
    hosts_data = st.session_state.hosts_data
    mentors_data = st.session_state.mentors_data
    students_data = st.session_state.students_data
    time_slots = st.session_state.time_slots

    if not hosts_data or not mentors_data or not students_data:
        st.session_state.solver_error = "Need at least 1 host, 1 mentor, and 1 student."
        st.session_state.schedule_result = None
        return
    if not time_slots:
        st.session_state.solver_error = "No time slots defined."
        st.session_state.schedule_result = None
        return

    hosts, mentors, students = _build_model_objects(hosts_data, mentors_data, students_data)

    with st.spinner("Solving ILP... this may take a moment"):
        try:
            result = solve(time_slots, hosts, mentors, students, verbose=False)
        except Exception as e:
            st.session_state.solver_error = f"Solver error: {e}"
            st.session_state.schedule_result = None
            return

    if result is None:
        st.session_state.solver_error = (
            "INFEASIBLE — no valid schedule exists under the given constraints. "
            "Check that every mentor has at least one student wanting their major "
            "with overlapping availability."
        )
        st.session_state.schedule_result = None
    else:
        st.session_state.schedule_result = result
        st.session_state.solver_error = None


def _render_schedule_tab(sessions: list[ScheduledSession]):
    days = _group_slots_by_day(st.session_state.time_slots)

    if len(days) > 1:
        day_tabs = st.tabs(list(days.keys()))
        for day_tab, (day_label, _) in zip(day_tabs, days.items()):
            with day_tab:
                day_sessions = [s for s in sessions if _parse_slot(s.time_slot)[0] == day_label]
                if not day_sessions:
                    st.info("No sessions scheduled for this day.")
                    continue
                df = _sessions_to_df(day_sessions)
                st.dataframe(df, width='stretch', hide_index=True)
    else:
        df = _sessions_to_df(sessions)
        st.dataframe(df, width='stretch', hide_index=True)


def _render_timetable_tab(sessions: list[ScheduledSession]):
    st.caption("Timetable view — separated by role and day. Cells show paired partners.")

    days = _group_slots_by_day(st.session_state.time_slots)

    role_tabs = st.tabs(["Hosts", "Mentors", "Students"])
    role_keys = ["host", "mentor", "student"]

    for role_tab, role_key in zip(role_tabs, role_keys):
        with role_tab:
            if len(days) > 1:
                day_tabs = st.tabs(list(days.keys()))
                for day_tab, (day_label, day_slots) in zip(day_tabs, days.items()):
                    with day_tab:
                        tt = _build_role_day_timetable(sessions, role_key, day_label, day_slots)
                        if tt.empty:
                            st.info("No sessions for this day.")
                        else:
                            st.dataframe(tt, width='stretch', hide_index=True)
            else:
                day_label = list(days.keys())[0]
                day_slots = days[day_label]
                tt = _build_role_day_timetable(sessions, role_key, day_label, day_slots)
                if tt.empty:
                    st.info("No sessions scheduled.")
                else:
                    st.dataframe(tt, width='stretch', hide_index=True)


def _render_summary_tab(
    sessions: list[ScheduledSession],
    mentors: list[Mentor],
    students: list[Student],
):
    col1, col2, col3 = st.columns(3)
    mentor_counts = Counter(s.mentor for s in sessions)
    student_covered = {s.student for s in sessions}

    with col1:
        st.metric("Total Sessions", len(sessions))
    with col2:
        st.metric("Mentors Active", f"{len(mentor_counts)} / {len(mentors)}")
    with col3:
        st.metric("Students Served", f"{len(student_covered)} / {len(students)}")

    unserved = [s.name for s in students if s.name not in student_covered]
    if unserved:
        st.warning(f"Students NOT served: {', '.join(unserved)}")

    st.subheader("Per-Mentor Breakdown")
    mentor_rows = []
    for m in mentors:
        cnt = mentor_counts.get(m.name, 0)
        mentor_rows.append({
            "Mentor": m.name,
            "Major": m.major,
            "Sessions": cnt,
            "Status": "✅" if cnt > 0 else "❌ Missing",
        })
    st.dataframe(pd.DataFrame(mentor_rows), width='stretch', hide_index=True)

    st.subheader("Per-Major Breakdown")
    major_counts = Counter(s.major for s in sessions)
    major_rows = [{"Major": maj, "Sessions": cnt} for maj, cnt in sorted(major_counts.items())]
    st.dataframe(pd.DataFrame(major_rows), width='stretch', hide_index=True)

    st.subheader("Constraint Verification")
    ok, msgs = _constraint_check(sessions, mentors)
    for msg in msgs:
        if "FAIL" in msg:
            st.error(msg)
        else:
            st.success(msg)


def _render_export_tab(sessions: list[ScheduledSession]):
    st.subheader("Download Results")
    st.caption("JSON includes input data + schedule + shift labels. Excel has per-role per-day timetable sheets.")

    col1, col2 = st.columns(2)

    with col1:
        json_bytes = _result_to_json_bytes(
            sessions,
            st.session_state.time_slots,
            st.session_state.hosts_data,
            st.session_state.mentors_data,
            st.session_state.students_data,
        )
        st.download_button(
            label="Download JSON",
            data=json_bytes,
            file_name="schedule_result.json",
            mime="application/json",
            width='stretch',
        )

    with col2:
        excel_bytes = _result_to_excel_bytes(sessions)
        st.download_button(
            label="Download Excel",
            data=excel_bytes,
            file_name="schedule_result.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            width='stretch',
        )

def main():
    st.title("Orientation Event Scheduler")
    st.caption(
        "Upload Excel files or enter data manually → edit on the fly → solve → export."
    )

    _render_sidebar()

    if not st.session_state.data_loaded:
        st.info(
            "**Get started** by uploading your Excel file(s) in the sidebar, "
            "or choose 'Manual entry' to type data directly."
        )
        with st.expander("Quick Start Guide"):
            st.markdown("""
**Supported Excel formats:**

1. **Combined workbook** — One `.xlsx` with tabs named `hosts`, `mentors`, `students`.
   Day columns contain text like `ca 1,2,3` or `5-12`.

2. **Separate files** — Three `.xlsx` files, one per role.
   - **Checkbox format**: Tabs named by date (e.g. `13/6`). TRUE/FALSE per shift.
   - **Text format**: Tabs named by role. Free-text day columns.

3. **Manual entry** — Define time slots in the sidebar, then add people in the data editor.

**Workflow:**
1. Upload Excel → data appears in the **Data Editor**, grouped by day
2. Toggle checkboxes, add/remove people, change majors
3. Go to **Shift Labels** to customize Ca 1-12 display names
4. Click **Run Solver** — see results split by day and role
5. Download as JSON or Excel
""")
        return
    main_editor, main_labels = st.tabs(["Data Editor", "Shift Labels"])

    with main_editor:
        _render_data_editor()

    with main_labels:
        _render_shift_label_editor()

    st.divider()
    _render_solver()


if __name__ == "__main__":
    main()
