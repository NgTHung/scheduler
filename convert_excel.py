"""
Excel → JSON converter for the Orientation Event Scheduler.

Supports TWO input formats, auto-detected by tab names:

Format A — "checkbox" (tabs named by date, e.g. "13/6", "14-6")
    3 separate workbooks, one per role. Each tab = one day.
    Hosts:            [id] [name]          [shift1 TRUE/FALSE] …
    Mentors/Students: [id] [major(merged)] [name] [shift1 TRUE/FALSE] …
    Row 1 = title (ignored), Row 2 = headers, Row 3+ = data.

Format B — "text" (tabs named by role, e.g. "mentors", "hosts")
    Single workbook OR 3 separate workbooks.
    Tabs: hosts / mentors / students.
    Mentors/Students: [major(merged)] [name] [day1_text] [day2_text] …
    Hosts:                            [name] [day1_text] [day2_text] …
    Day columns contain text like "ca 1,2,3", "Không", "5-12", etc.
    Header row auto-detected (contains date-like column labels).

Usage
-----
    # 3 separate workbooks (checkbox or text — auto-detected)
    python convert_excel.py --hosts hosts.xlsx -m mentors.xlsx -s students.xlsx

    # Single combined workbook (tabs = roles)
    python convert_excel.py -i combined.xlsx

    # Output path (default: data.json)
    python convert_excel.py -i combined.xlsx -o schedule.json
"""

from __future__ import annotations

import argparse
import datetime as _dt
import json
import re
import sys
from pathlib import Path
from typing import Any

try:
    import openpyxl
    from openpyxl.worksheet.worksheet import Worksheet
except ImportError:
    print("openpyxl is required.  Install it with:  pip install openpyxl")
    sys.exit(1)

ROLE_TAB_ALIASES: dict[str, str] = {
    "host": "host", "hosts": "host",
    "mentor": "mentor", "mentors": "mentor",
    "student": "student", "students": "student",
}


def _normalize_day_label(raw: str) -> str:
    """
    Normalize a day label so that different sources produce identical slot IDs.

    Rules:
      • Strip whitespace
      • Replace  -  .  with  /      ("13-6" → "13/6",  "13.06" → "13/06")
      • Remove leading zeros in each part  ("14/06" → "14/6")

    Examples:
        "13-6"   → "13/6"
        "14-06"  → "14/6"
        "13/6"   → "13/6"   (no change)
        "14/06"  → "14/6"
    """
    s = raw.strip().replace("-", "/").replace(".", "/")
    # Remove leading zeros: "14/06" → "14/6"
    parts = s.split("/")
    parts = [p.lstrip("0") or "0" for p in parts]
    return "/".join(parts)

def _get_merged_cell_value(ws: Worksheet, row: int, col: int):
    """Return the visual value of a cell that might be inside a merged range."""
    cell = ws.cell(row=row, column=col)
    if cell.value is not None:
        return cell.value
    for merged_range in ws.merged_cells.ranges:
        if cell.coordinate in merged_range:
            return ws.cell(
                row=merged_range.min_row,
                column=merged_range.min_col,
            ).value
    return None

_SHIFT_PREFIX_RE = re.compile(r"(?i)\b(dạ\s+)?ca\s*", re.UNICODE)
_NO_AVAILABILITY = {"không", "khong", "no", "none", "n/a", "x", "-", ""}


def parse_shift_text(raw: Any) -> set[int]:
    """
    Parse a free-text availability cell into a set of shift numbers.

    Handles all observed patterns:
        "ca 9"              → {9}
        "2, 6, 11, 12"     → {2, 6, 11, 12}
        "Ca 3,5,6,7,8"     → {3, 5, 6, 7, 8}
        "9 - 10 - 11 - 12" → {9, 10, 11, 12}
        "5-12"              → {5, 6, 7, 8, 9, 10, 11, 12}
        "ca 4, 5-9"         → {4, 5, 6, 7, 8, 9}
        "1;2;3"             → {1, 2, 3}
        "Dạ ca 2, ca 3"     → {2, 3}
        "Ca 7, ca 8"        → {7, 8}
        "ca 11, 12"         → {11, 12}
        "Không"             → {}
        ""                  → {}
    """
    if raw is None:
        return set()
    # Excel may auto-convert shift text like "1,11,12" into a date.
    # Recover candidate shift numbers from day, month, and 2-digit year.
    # Only include the year component when it looks like an explicit part of
    # the input (small 2-digit number) rather than an auto-filled current year.
    if isinstance(raw, (_dt.datetime, _dt.date)):
        candidates = {raw.day, raw.month}
        y = raw.year % 100
        if 1 <= y <= 20:           # e.g. 2012 → 12 (plausible shift)
            candidates.add(y)
        return candidates
    if isinstance(raw, int):
        return {int(raw)}
    if isinstance(raw, float):
        if raw == int(raw):          # e.g. 3.0 → shift 3
            return {int(raw)}
        # e.g. 11.12 → treat as text "11.12" (shifts 11 and 12)

    s = str(raw).strip()
    if s.lower() in _NO_AVAILABILITY:
        return set()

    # Strip Vietnamese "ca" / "Dạ ca" prefixes (all occurrences)
    s = _SHIFT_PREFIX_RE.sub(" ", s)

    shifts: set[int] = set()

    # Split by comma, semicolon, or dot first (preserves ranges like "5-12")
    tokens = re.split(r"[,;.]+", s)
    for token in tokens:
        token = token.strip()
        if not token:
            continue

        # Check for range: "5-12", "5 - 12"
        range_match = re.match(r"^(\d+)\s*-\s*(\d+)$", token)
        if range_match:
            lo, hi = int(range_match.group(1)), int(range_match.group(2))
            shifts.update(range(lo, hi + 1))
            continue

        # Check if token is "9 - 10 - 11" (dash-separated list)
        if "-" in token:
            parts = re.split(r"\s*-\s*", token)
            for part in parts:
                part = part.strip()
                if part.isdigit():
                    shifts.add(int(part))
            continue

        # Plain number(s) — extract all digit sequences
        for num_str in re.findall(r"\d+", token):
            shifts.add(int(num_str))

    return shifts

def _tab_to_role(tab_name: str) -> str | None:
    """Map a tab name to a canonical role, or None."""
    return ROLE_TAB_ALIASES.get(tab_name.strip().lower())


def detect_workbook_format(wb) -> str:
    """Return 'text' if any tab is role-named, else 'checkbox'."""
    for name in wb.sheetnames:
        if _tab_to_role(name) is not None:
            return "text"
    return "checkbox"


def _looks_like_date(val: Any) -> bool:
    """Does this value look like a date column header? (e.g. '13/6', '14-06', or a datetime object)"""
    if val is None:
        return False
    if isinstance(val, (_dt.datetime, _dt.date)):
        return True
    s = str(val).strip()
    return bool(re.search(r"\d+\s*[/\-\.]\s*\d+", s))


def _header_to_day_label(val: Any) -> str:
    """Convert a header cell value to a day label string (day/month)."""
    if isinstance(val, (_dt.datetime, _dt.date)):
        return f"{val.day}/{val.month}"
    return str(val).strip()


def _find_text_layout(ws: Worksheet, role: str) -> dict:
    """
    Auto-detect header row and column layout for a text-format sheet.

    Returns dict with keys:
        header_row, data_start, name_col, major_col (or None),
        day_cols [(col_index, day_label), ...]
    """
    for try_row in range(1, 5):
        day_cols: list[tuple[int, str]] = []
        for col in range(1, min((ws.max_column or 20) + 1, 50)):
            val = ws.cell(row=try_row, column=col).value
            if _looks_like_date(val):
                day_cols.append((col, _header_to_day_label(val)))

        if day_cols:
            # Normalize day labels so they match across formats
            day_cols = [(c, _normalize_day_label(lbl)) for c, lbl in day_cols]
            first_day_col = day_cols[0][0]
            name_col = first_day_col - 1

            major_col = None
            if role != "host":
                candidate = first_day_col - 2
                if candidate >= 1:
                    major_col = candidate

            return {
                "header_row": try_row,
                "data_start": try_row + 1,
                "name_col": max(name_col, 1),
                "major_col": major_col,
                "day_cols": day_cols,
            }

    raise ValueError(
        "Could not find date-like column headers in the first 4 rows. "
        "Expected columns like '13/6' or '14-06'."
    )

def _detect_shift_count(ws: Worksheet, header_row: int, start_col: int) -> int:
    count = 0
    col = start_col
    while True:
        val = ws.cell(row=header_row, column=col).value
        if val is None:
            break
        count += 1
        col += 1
    return count


def _is_truthy(val: Any) -> bool:
    if val is None:
        return False
    if isinstance(val, bool):
        return val
    if isinstance(val, (int, float)):
        return val != 0
    s = str(val).strip().upper()
    return s in ("TRUE", "1", "YES", "Y", "✓", "☑", "X")


def _parse_checkbox_hosts(ws: Worksheet, day_label_raw: str):
    day_label = _normalize_day_label(day_label_raw)
    HEADER_ROW, DATA_START = 2, 3
    NAME_COL, SHIFT_START = 2, 3

    n = _detect_shift_count(ws, HEADER_ROW, SHIFT_START)
    if n == 0:
        return [], []

    slots = [f"{day_label}_{i}" for i in range(1, n + 1)]
    entries: list[dict] = []
    row = DATA_START
    while True:
        name = ws.cell(row=row, column=NAME_COL).value
        if name is None or str(name).strip() == "":
            break
        name = str(name).strip()
        avail = [slots[i] for i in range(n)
                 if _is_truthy(ws.cell(row=row, column=SHIFT_START + i).value)]
        entries.append({"name": name, "available_slots": avail})
        row += 1
    return entries, slots


def _parse_checkbox_role(ws: Worksheet, day_label_raw: str, role: str):
    day_label = _normalize_day_label(day_label_raw)
    HEADER_ROW, DATA_START = 2, 3
    MAJOR_COL, NAME_COL, SHIFT_START = 2, 3, 4

    n = _detect_shift_count(ws, HEADER_ROW, SHIFT_START)
    if n == 0:
        return [], []

    slots = [f"{day_label}_{i}" for i in range(1, n + 1)]
    entries: list[dict] = []
    row = DATA_START
    while True:
        name = ws.cell(row=row, column=NAME_COL).value
        if name is None or str(name).strip() == "":
            peek_empty = all(
                ws.cell(row=row + k, column=NAME_COL).value in (None, "")
                for k in range(1, 4)
            )
            if peek_empty:
                break
            row += 1
            continue

        name = str(name).strip()
        major_raw = _get_merged_cell_value(ws, row, MAJOR_COL)
        major = str(major_raw).strip() if major_raw else "UNKNOWN"
        avail = [slots[i] for i in range(n)
                 if _is_truthy(ws.cell(row=row, column=SHIFT_START + i).value)]

        key = "major" if role == "mentor" else "desired_major"
        entries.append({"name": name, key: major, "available_slots": avail})
        row += 1
    return entries, slots

def _parse_text_sheet(ws: Worksheet, role: str):
    """
    Parse a text-format sheet where days are columns and availability
    is free text like "ca 1,2,3".

    Returns (entries, all_slots, n_shifts).
    """
    layout = _find_text_layout(ws, role)
    data_start = layout["data_start"]
    name_col = layout["name_col"]
    major_col = layout["major_col"]
    day_cols = layout["day_cols"]

    entries: list[dict] = []
    global_max_shift = 0

    row = data_start
    while True:
        name = ws.cell(row=row, column=name_col).value
        if name is None or str(name).strip() == "":
            peek_empty = all(
                ws.cell(row=row + k, column=name_col).value in (None, "")
                for k in range(1, 6)
            )
            if peek_empty:
                break
            row += 1
            continue

        name = str(name).strip()

        # Major (for mentors / students)
        major = None
        if major_col is not None and role != "host":
            major_raw = _get_merged_cell_value(ws, row, major_col)
            major = str(major_raw).strip() if major_raw else "UNKNOWN"

        # Parse availability from each day column
        available: list[str] = []
        for col_idx, day_label in day_cols:
            cell_val = ws.cell(row=row, column=col_idx).value
            shift_nums = parse_shift_text(cell_val)
            if shift_nums:
                global_max_shift = max(global_max_shift, max(shift_nums))
            available.extend(f"{day_label}_{s}" for s in sorted(shift_nums))

        entry: dict[str, Any] = {"name": name, "available_slots": available}
        if role == "mentor":
            entry["major"] = major or "UNKNOWN"
        elif role == "student":
            entry["desired_major"] = major or "UNKNOWN"

        entries.append(entry)
        row += 1

    # Generate complete time-slots list (all days × shifts 1..N)
    all_slots: list[str] = []
    for _, day_label in day_cols:
        for i in range(1, global_max_shift + 1):
            all_slots.append(f"{day_label}_{i}")

    return entries, all_slots, global_max_shift

def parse_workbook(path: str | Path, role: str):
    """
    Parse a workbook for a given role, auto-detecting format per file.
    Returns (entries_list, time_slots).
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    fmt = detect_workbook_format(wb)
    fname = Path(path).name

    all_entries: dict[str, dict] = {}
    all_slots: list[str] = []

    if fmt == "text":
        # Find the tab matching this role
        target_ws = None
        for name in wb.sheetnames:
            if _tab_to_role(name) == role:
                target_ws = wb[name]
                break
        if target_ws is None:
            target_ws = wb.active

        print(f"  Parsing '{target_ws.title}' in {fname} (text format) ...")
        entries, slots, n_shifts = _parse_text_sheet(target_ws, role)
        all_slots = slots
        for entry in entries:
            all_entries[entry["name"]] = entry
        print(f"  Detected {n_shifts} shifts/day")

    else:  # checkbox
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            day_label = str(sheet_name).strip()
            print(f"  Parsing tab '{day_label}' in {fname} (checkbox format) ...")

            if role == "host":
                entries, slots = _parse_checkbox_hosts(ws, day_label)
            else:
                entries, slots = _parse_checkbox_role(ws, day_label, role)

            all_slots.extend(slots)
            for entry in entries:
                name = entry["name"]
                if name in all_entries:
                    all_entries[name]["available_slots"].extend(
                        entry["available_slots"]
                    )
                else:
                    all_entries[name] = entry

    wb.close()
    return list(all_entries.values()), all_slots


def parse_combined_workbook(path: str | Path):
    """
    Parse a single workbook with role-named tabs (hosts, mentors, students).
    Returns (hosts, mentors, students, all_slots).
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    fname = Path(path).name

    result: dict[str, tuple[list[dict], list[str]]] = {}
    global_max_shift = 0

    for sheet_name in wb.sheetnames:
        role = _tab_to_role(sheet_name)
        if role is None:
            print(f"  Skipping unrecognized tab '{sheet_name}' in {fname}")
            continue

        ws = wb[sheet_name]
        print(f"  Parsing tab '{sheet_name}' → role={role} in {fname} (text format) ...")
        entries, slots, n_shifts = _parse_text_sheet(ws, role)
        global_max_shift = max(global_max_shift, n_shifts)
        result[role] = (entries, slots)

    wb.close()

    hosts = result.get("host", ([], []))
    mentors = result.get("mentor", ([], []))
    students = result.get("student", ([], []))

    # Unify time-slots across all roles
    seen: set[str] = set()
    all_slots: list[str] = []
    for _, slots in [hosts, mentors, students]:
        for s in slots:
            if s not in seen:
                seen.add(s)
                all_slots.append(s)

    return hosts[0], mentors[0], students[0], all_slots

def convert(
    hosts_path: str | None = None,
    mentors_path: str | None = None,
    students_path: str | None = None,
    combined_path: str | None = None,
    output_path: str = "data.json",
    slot_mapping: dict[str, str] | None = None,
):
    print(f"\n{'=' * 60}")
    print("  Excel → JSON Converter")
    print(f"{'=' * 60}\n")

    hosts: list[dict] = []
    mentors: list[dict] = []
    students: list[dict] = []
    slot_lists: list[list[str]] = []

    
    combined_roles: dict[str, tuple[list[dict], list[str]]] = {}
    if combined_path:
        wb = openpyxl.load_workbook(combined_path, data_only=True)
        fname = Path(combined_path).name
        for sheet_name in wb.sheetnames:
            role = _tab_to_role(sheet_name)
            if role is None:
                continue
            ws = wb[sheet_name]
            print(f"  Parsing tab '{sheet_name}' → role={role} in {fname} (text format) ...")
            entries, slots, _ = _parse_text_sheet(ws, role)
            combined_roles[role] = (entries, slots)
        wb.close()

    
    role_sources = [
        ("host",    hosts_path,    "Hosts"),
        ("mentor",  mentors_path,  "Mentors"),
        ("student", students_path, "Students"),
    ]

    for role, override_path, label in role_sources:
        if override_path:
            src = f"separate file ({Path(override_path).name})"
            print(f"  [{label}] from {src}")
            entries, slots = parse_workbook(override_path, role)
        elif role in combined_roles:
            src = f"combined workbook"
            print(f"  [{label}] from {src}")
            entries, slots = combined_roles[role]
        else:
            print(f"  [{label}] ⚠  NO SOURCE — skipped")
            entries, slots = [], []

        print(f"       → {len(entries)} {label.lower()}")
        slot_lists.append(slots)

        if role == "host":
            hosts = entries
        elif role == "mentor":
            mentors = entries
        else:
            students = entries

    print()

    
    seen: set[str] = set()
    all_slots: list[str] = []
    for slot_list in slot_lists:
        for s in slot_list:
            if s not in seen:
                seen.add(s)
                all_slots.append(s)

    
    if slot_mapping:
        all_slots = [slot_mapping.get(s, s) for s in all_slots]
        for person_list in (hosts, mentors, students):
            for person in person_list:
                person["available_slots"] = [
                    slot_mapping.get(s, s) for s in person["available_slots"]
                ]

    
    output = {
        "time_slots": all_slots,
        "hosts": hosts,
        "mentors": mentors,
        "students": students,
    }

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(output, f, indent=2, ensure_ascii=False)

    print(f"{'=' * 60}")
    print(f"  Written to : {output_path}")
    print(f"  Time-slots : {len(all_slots)}")
    print(f"  Hosts      : {len(hosts)}")
    print(f"  Mentors    : {len(mentors)}")
    print(f"  Students   : {len(students)}")

    majors = sorted({m.get("major", "?") for m in mentors})
    print(f"  Majors     : {', '.join(majors)}")

    for label, people in [("Hosts", hosts), ("Mentors", mentors), ("Students", students)]:
        no_avail = [p["name"] for p in people if not p["available_slots"]]
        if no_avail:
            print(f"\n  ⚠ {label} with ZERO availability: {', '.join(no_avail)}")

    print(f"{'=' * 60}\n")
    return output

def main():
    parser = argparse.ArgumentParser(
        description="Convert Excel availability sheets to scheduler JSON.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Formats (auto-detected by tab names):

  Format A — "checkbox" (3 workbooks, tabs = dates)
    Hosts:    [id] [name]          [shift1 TRUE/FALSE] …
    Mentors:  [id] [major(merged)] [name] [shift1 TRUE/FALSE] …
    Students: [id] [major(merged)] [name] [shift1 TRUE/FALSE] …

  Format B — "text" (1 workbook, tabs = roles OR 3 workbooks with role tabs)
    Mentors:  [major(merged)] [name] [day1_text] [day2_text] …
    Students: [major(merged)] [name] [day1_text] [day2_text] …
    Hosts:                    [name] [day1_text] [day2_text] …
    Day text: "ca 1,2,3", "5-12", "Không", etc.

  Mix & match: use -i for base, override any role with --hosts/-m/-s.
  Each file auto-detects its own format independently.

Examples:
  python convert_excel.py -i combined.xlsx
  python convert_excel.py --hosts hosts.xlsx -m mentors.xlsx -s students.xlsx
  python convert_excel.py -i combined.xlsx --hosts hosts_checkbox.xlsx
  python convert_excel.py -i data.xlsx -o schedule.json
""",
    )
    parser.add_argument("-i", "--input", dest="combined_file",
                        help="Combined workbook with tabs: hosts, mentors, students")
    parser.add_argument("--hosts", dest="hosts_file",
                        help="Hosts workbook (overrides -i for hosts)")
    parser.add_argument("-m", "--mentors", dest="mentors_file",
                        help="Mentors workbook (overrides -i for mentors)")
    parser.add_argument("-s", "--students", dest="students_file",
                        help="Students workbook (overrides -i for students)")
    parser.add_argument("-o", "--output", default="data.json",
                        help="Output JSON path (default: data.json)")
    parser.add_argument("--slot-map", dest="slot_map_file",
                        help="JSON file mapping slot IDs to readable labels")

    args = parser.parse_args()

    
    has_combined = bool(args.combined_file)
    has_separate = bool(args.hosts_file or args.mentors_file or args.students_file)

    if has_combined or has_separate:
        # Validate all supplied paths exist
        for label, path in [("Combined", args.combined_file),
                            ("Hosts", args.hosts_file),
                            ("Mentors", args.mentors_file),
                            ("Students", args.students_file)]:
            if path and not Path(path).exists():
                print(f"ERROR: {label} file not found: {path}")
                sys.exit(1)

        # In 3-file mode without -i, prompt for any missing role files
        if not has_combined:
            if not args.hosts_file:
                args.hosts_file = input("Path to hosts Excel file: ").strip().strip('"')
            if not args.mentors_file:
                args.mentors_file = input("Path to mentors Excel file: ").strip().strip('"')
            if not args.students_file:
                args.students_file = input("Path to students Excel file: ").strip().strip('"')

        slot_mapping = None
        if args.slot_map_file:
            with open(args.slot_map_file, encoding="utf-8") as f:
                slot_mapping = json.load(f)

        convert(
            hosts_path=args.hosts_file,
            mentors_path=args.mentors_file,
            students_path=args.students_file,
            combined_path=args.combined_file,
            output_path=args.output,
            slot_mapping=slot_mapping,
        )
    else:
        # Fully interactive
        print("No files specified. Choose a mode:\n")
        print("  1) Single workbook (tabs = hosts, mentors, students)")
        print("  2) Three separate workbooks")
        print("  3) Hybrid (combined base + per-role overrides)")
        choice = input("\nChoice [1/2/3]: ").strip()

        if choice == "1":
            path = input("Path to workbook: ").strip().strip('"')
            if not Path(path).exists():
                print(f"ERROR: File not found: {path}")
                sys.exit(1)
            convert(combined_path=path, output_path=args.output)
        elif choice == "3":
            cp = input("Path to combined workbook: ").strip().strip('"')
            if not Path(cp).exists():
                print(f"ERROR: File not found: {cp}")
                sys.exit(1)
            print("\nFor each role, enter a separate file path to override,")
            print("or press Enter to use the combined workbook.\n")
            hp = input("  Hosts override (Enter=skip): ").strip().strip('"') or None
            mp = input("  Mentors override (Enter=skip): ").strip().strip('"') or None
            sp = input("  Students override (Enter=skip): ").strip().strip('"') or None
            convert(hosts_path=hp, mentors_path=mp, students_path=sp,
                    combined_path=cp, output_path=args.output)
        else:
            hp = input("Path to hosts Excel file: ").strip().strip('"')
            mp = input("Path to mentors Excel file: ").strip().strip('"')
            sp = input("Path to students Excel file: ").strip().strip('"')
            for lab, p in [("Hosts", hp), ("Mentors", mp), ("Students", sp)]:
                if not Path(p).exists():
                    print(f"ERROR: {lab} file not found: {p}")
                    sys.exit(1)
            convert(hosts_path=hp, mentors_path=mp, students_path=sp,
                    output_path=args.output)


if __name__ == "__main__":
    main()
