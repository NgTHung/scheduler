"""
Generate a sample COMBINED Excel workbook in text format (Format B)
to test the converter with messy Vietnamese shift text.

Creates: sample_combined.xlsx
  Tabs: "mentors", "hosts", "students"
  Layout: [major(merged)] [name] [13/6] [14/06]
"""

import openpyxl


def create_combined():
    wb = openpyxl.Workbook()

    # ────────────────────────── MENTORS ──────────────────────────── #
    ws = wb.active
    ws.title = "mentors"

    ws.cell(1, 1, "Ngành")
    ws.cell(1, 2, "Họ tên")
    ws.cell(1, 3, "13/6")
    ws.cell(1, 4, "14/06")

    mentors = [
        # (name, major, day1_text, day2_text)
        ("Huỳnh Ngọc Thanh Tâm",  "Data Analytics", "ca 9",                       None),
        ("Lê Vũ Ngọc Giang",      "Data Analytics", "2, 6, 11, 12",               "2, 6, 11, 12"),
        ("Nông Thị Ngân Thương",   "Data Analytics", "Không",                      "Dạ ca 2, ca 3"),
        ("Lê Thị Ngọc Anh",       "Data Analytics", "1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12", "1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12"),
        ("Nguyễn Lê Trâm Anh",    "Data Analytics", "1;2;3",                      "1;2;3"),

        ("Nguyễn Ngọc Tường Vi",   "Finance",        "Không",                      "10,11,12"),
        ("Lê Bùi Trần Thái Hồng", "Finance",        "Ca 1,2,3,4,6,7,8,11,12",    "Ca 7,8,11,12"),
        ("Đặng Gia Minh",         "Finance",        "9",                           "Không"),
        ("Nguyễn Huỳnh Nga",      "Finance",        "4, 10",                      "4, 10"),
        ("Đoàn Tuấn Minh",        "Finance",        "3, 4, 5, 6, 7, 8, 9, 10, 11, 12", "3, 4, 5, 6, 7, 8, 9, 10, 11, 12"),

        ("Eve_M",                  "Sales",          "Ca 8",                       "Ca 4"),
        ("Frank_M",                "Sales",          "ca 10",                      "Không"),

        ("Grace_M",                "Marketing",      "Không",                      "6;7;8"),
    ]

    row = 2
    major_start: dict[str, int] = {}
    major_count: dict[str, int] = {}

    for name, major, d1, d2 in mentors:
        ws.cell(row, 2, name)
        ws.cell(row, 3, d1)
        ws.cell(row, 4, d2)

        if major not in major_start:
            major_start[major] = row
            major_count[major] = 1
            ws.cell(row, 1, major)
        else:
            major_count[major] += 1

        row += 1

    for major, start in major_start.items():
        cnt = major_count[major]
        if cnt > 1:
            ws.merge_cells(start_row=start, start_column=1,
                           end_row=start + cnt - 1, end_column=1)

    # ────────────────────────── HOSTS ────────────────────────────── #
    ws2 = wb.create_sheet("hosts")
    ws2.cell(1, 1, "Họ tên")
    ws2.cell(1, 2, "13/6")
    ws2.cell(1, 3, "14/06")

    hosts = [
        ("Alice_H",   "1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12", "1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12"),
        ("Bob_H",     "Ca 1,2,3,4,5,6",                          "Ca 7,8,9,10,11,12"),
        ("Charlie_H", "ca 5-12",                                  "ca 1-6"),
    ]

    for r, (name, d1, d2) in enumerate(hosts, start=2):
        ws2.cell(r, 1, name)
        ws2.cell(r, 2, d1)
        ws2.cell(r, 3, d2)

    # ────────────────────────── STUDENTS ─────────────────────────── #
    ws3 = wb.create_sheet("students")
    ws3.cell(1, 1, "Ngành")
    ws3.cell(1, 2, "Họ tên")
    ws3.cell(1, 3, "13/6")
    ws3.cell(1, 4, "14/06")

    students = [
        ("S1_Liam",    "Data Analytics", "Ca 1,2,3",     "Ca 4,5"),
        ("S2_Noah",    "Data Analytics", "ca 9",         "ca 2, 3"),
        ("S3_Emma",    "Finance",        "4, 10",        "4, 10"),
        ("S4_Olivia",  "Finance",        "Ca 3,5,6,7,8", "Ca 7,8,11,12"),
        ("S5_Ava",     "Finance",        "9 - 10 - 11 - 12", "5 - 6 - 12"),
        ("S6_Sophia",  "Sales",          "Ca 8",          "Ca 4"),
        ("S7_Mia",     "Sales",          "ca 10",         "Không"),
        ("S8_James",   "Marketing",      "Không",         "6;7;8"),
    ]

    row = 2
    ms: dict[str, int] = {}
    mc: dict[str, int] = {}
    for name, major, d1, d2 in students:
        ws3.cell(row, 2, name)
        ws3.cell(row, 3, d1)
        ws3.cell(row, 4, d2)
        if major not in ms:
            ms[major] = row
            mc[major] = 1
            ws3.cell(row, 1, major)
        else:
            mc[major] += 1
        row += 1

    for major, start in ms.items():
        cnt = mc[major]
        if cnt > 1:
            ws3.merge_cells(start_row=start, start_column=1,
                            end_row=start + cnt - 1, end_column=1)

    wb.save("sample_combined.xlsx")
    print("Created sample_combined.xlsx (text format, 3 role tabs)")


if __name__ == "__main__":
    create_combined()
