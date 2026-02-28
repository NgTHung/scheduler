"""
Generate sample Excel files to test the convert_excel.py converter.

Creates 3 workbooks: sample_hosts.xlsx, sample_mentors.xlsx, sample_students.xlsx
Each with 2 day-tabs ("13/6" and "14/6"), 6 shifts per day.
"""

import openpyxl
from openpyxl.utils import get_column_letter


def create_hosts():
    wb = openpyxl.Workbook()

    # ── Day 1: "13/6" ──
    ws = wb.active
    ws.title = "13-6"

    ws.cell(1, 1, "HOST AVAILABILITY")
    ws.cell(2, 1, "ID")
    ws.cell(2, 2, "Name")
    for i in range(1, 7):
        ws.cell(2, 2 + i, i)

    hosts_data = [
        ("Alice_H",   [True,  True,  True,  True,  True,  True]),
        ("Bob_H",     [True,  True,  False, True,  True,  False]),
    ]
    for r, (name, shifts) in enumerate(hosts_data, start=3):
        ws.cell(r, 1, r - 2)
        ws.cell(r, 2, name)
        for i, val in enumerate(shifts):
            ws.cell(r, 3 + i, val)

    # ── Day 2: "14/6" ──
    ws2 = wb.create_sheet("14-6")
    ws2.cell(1, 1, "HOST AVAILABILITY")
    ws2.cell(2, 1, "ID")
    ws2.cell(2, 2, "Name")
    for i in range(1, 7):
        ws2.cell(2, 2 + i, i)

    hosts_data_2 = [
        ("Alice_H",   [True,  False, True,  True,  False, True]),
        ("Bob_H",     [False, True,  True,  False, True,  True]),
    ]
    for r, (name, shifts) in enumerate(hosts_data_2, start=3):
        ws2.cell(r, 1, r - 2)
        ws2.cell(r, 2, name)
        for i, val in enumerate(shifts):
            ws2.cell(r, 3 + i, val)

    wb.save("sample_hosts.xlsx")
    print("Created sample_hosts.xlsx")


def create_mentors():
    wb = openpyxl.Workbook()

    # ── Day 1: "13/6" ──
    ws = wb.active
    ws.title = "13-6"

    ws.cell(1, 1, "MENTOR AVAILABILITY")
    ws.cell(2, 1, "ID")
    ws.cell(2, 2, "Major")
    ws.cell(2, 3, "Name")
    for i in range(1, 7):
        ws.cell(2, 3 + i, i)

    mentors = [
        # (name, major, shifts)
        ("Carol_M",  "HR",        [True,  True,  True,  False, False, False]),
        ("Dave_M",   "HR",        [False, False, False, True,  True,  False]),
        ("Eve_M",    "Sales",     [True,  True,  False, True,  False, False]),
        ("Frank_M",  "Sales",     [False, True,  True,  False, False, True]),
        ("Grace_M",  "Marketing", [True,  False, True,  True,  False, True]),
    ]

    row = 3
    major_start: dict[str, int] = {}
    major_count: dict[str, int] = {}

    for idx, (name, major, shifts) in enumerate(mentors):
        ws.cell(row, 1, idx + 1)
        ws.cell(row, 3, name)
        for i, val in enumerate(shifts):
            ws.cell(row, 4 + i, val)

        if major not in major_start:
            major_start[major] = row
            major_count[major] = 1
            ws.cell(row, 2, major)
        else:
            major_count[major] += 1

        row += 1

    # Merge major cells
    for major, start in major_start.items():
        cnt = major_count[major]
        if cnt > 1:
            ws.merge_cells(
                start_row=start, start_column=2,
                end_row=start + cnt - 1, end_column=2,
            )

    # ── Day 2: "14/6" ── (same layout, different availability)
    ws2 = wb.create_sheet("14-6")
    ws2.cell(1, 1, "MENTOR AVAILABILITY")
    ws2.cell(2, 1, "ID")
    ws2.cell(2, 2, "Major")
    ws2.cell(2, 3, "Name")
    for i in range(1, 7):
        ws2.cell(2, 3 + i, i)

    mentors2 = [
        ("Carol_M",  "HR",        [False, True,  False, True,  True,  False]),
        ("Dave_M",   "HR",        [True,  False, True,  False, False, True]),
        ("Eve_M",    "Sales",     [False, False, True,  True,  True,  False]),
        ("Frank_M",  "Sales",     [True,  True,  False, False, True,  True]),
        ("Grace_M",  "Marketing", [False, True,  True,  False, True,  False]),
    ]

    row = 3
    major_start2: dict[str, int] = {}
    major_count2: dict[str, int] = {}

    for idx, (name, major, shifts) in enumerate(mentors2):
        ws2.cell(row, 1, idx + 1)
        ws2.cell(row, 3, name)
        for i, val in enumerate(shifts):
            ws2.cell(row, 4 + i, val)

        if major not in major_start2:
            major_start2[major] = row
            major_count2[major] = 1
            ws2.cell(row, 2, major)
        else:
            major_count2[major] += 1

        row += 1

    for major, start in major_start2.items():
        cnt = major_count2[major]
        if cnt > 1:
            ws2.merge_cells(
                start_row=start, start_column=2,
                end_row=start + cnt - 1, end_column=2,
            )

    wb.save("sample_mentors.xlsx")
    print("Created sample_mentors.xlsx")


def create_students():
    wb = openpyxl.Workbook()

    # ── Day 1: "13/6" ──
    ws = wb.active
    ws.title = "13-6"

    ws.cell(1, 1, "STUDENT AVAILABILITY")
    ws.cell(2, 1, "ID")
    ws.cell(2, 2, "Major")
    ws.cell(2, 3, "Name")
    for i in range(1, 7):
        ws.cell(2, 3 + i, i)

    students = [
        ("S1_Liam",   "HR",        [True,  True,  False, False, False, False]),
        ("S2_Noah",   "HR",        [False, True,  False, True,  True,  False]),
        ("S3_Emma",   "HR",        [False, False, True,  False, False, False]),
        ("S4_Olivia", "Sales",     [True,  True,  True,  False, False, False]),
        ("S5_Ava",    "Sales",     [False, False, True,  False, False, True]),
        ("S6_Sophia", "Sales",     [False, False, False, True,  True,  False]),
        ("S7_Mia",    "Marketing", [True,  False, True,  False, False, False]),
        ("S8_James",  "Marketing", [False, False, False, True,  True,  False]),
    ]

    row = 3
    major_start: dict[str, int] = {}
    major_count: dict[str, int] = {}

    for idx, (name, major, shifts) in enumerate(students):
        ws.cell(row, 1, idx + 1)
        ws.cell(row, 3, name)
        for i, val in enumerate(shifts):
            ws.cell(row, 4 + i, val)

        if major not in major_start:
            major_start[major] = row
            major_count[major] = 1
            ws.cell(row, 2, major)
        else:
            major_count[major] += 1

        row += 1

    for major, start in major_start.items():
        cnt = major_count[major]
        if cnt > 1:
            ws.merge_cells(
                start_row=start, start_column=2,
                end_row=start + cnt - 1, end_column=2,
            )

    # ── Day 2: "14/6" ──
    ws2 = wb.create_sheet("14-6")
    ws2.cell(1, 1, "STUDENT AVAILABILITY")
    ws2.cell(2, 1, "ID")
    ws2.cell(2, 2, "Major")
    ws2.cell(2, 3, "Name")
    for i in range(1, 7):
        ws2.cell(2, 3 + i, i)

    students2 = [
        ("S1_Liam",   "HR",        [False, False, True,  True,  False, False]),
        ("S2_Noah",   "HR",        [True,  False, False, False, True,  True]),
        ("S3_Emma",   "HR",        [False, True,  True,  False, False, False]),
        ("S4_Olivia", "Sales",     [False, True,  False, True,  False, True]),
        ("S5_Ava",    "Sales",     [True,  False, False, False, True,  False]),
        ("S6_Sophia", "Sales",     [False, False, True,  True,  False, True]),
        ("S7_Mia",    "Marketing", [False, True,  False, True,  True,  False]),
        ("S8_James",  "Marketing", [True,  False, True,  False, False, True]),
    ]

    row = 3
    major_start2: dict[str, int] = {}
    major_count2: dict[str, int] = {}

    for idx, (name, major, shifts) in enumerate(students2):
        ws2.cell(row, 1, idx + 1)
        ws2.cell(row, 3, name)
        for i, val in enumerate(shifts):
            ws2.cell(row, 4 + i, val)

        if major not in major_start2:
            major_start2[major] = row
            major_count2[major] = 1
            ws2.cell(row, 2, major)
        else:
            major_count2[major] += 1

        row += 1

    for major, start in major_start2.items():
        cnt = major_count2[major]
        if cnt > 1:
            ws2.merge_cells(
                start_row=start, start_column=2,
                end_row=start + cnt - 1, end_column=2,
            )

    wb.save("sample_students.xlsx")
    print("Created sample_students.xlsx")


if __name__ == "__main__":
    create_hosts()
    create_mentors()
    create_students()
    print("\nDone! Now run:")
    print("  python convert_excel.py --hosts sample_hosts.xlsx -m sample_mentors.xlsx -s sample_students.xlsx -o data.json")
    print("  python main.py data.json")
